"""PDF ve Excel dosyalarini `AyristirilmisBelge`'ye cevirir.

TASARIM KARARI - NEDEN "HATA FIRLATMA" YERINE "UYARI TOPLAMA"
-------------------------------------------------------------
Gercek dunyada finansal belgeler kusurludur: 40 sayfalik bir faaliyet
raporunun 3 sayfasi taranmis goruntu olabilir, bir Excel'in ikinci sayfasi
bos olabilir, bir tablo hucresi birlestirilmis olabilir. Bunlarin her birinde
istisna firlatmak, kullanicinin okunabilir 37 sayfayi da KAYBETMESI demektir.

Bu yuzden ayristirici yalnizca "dosya hic acilamadi" durumunda firlatir;
geri kalan her sorun `uyarilar` listesine yazilir ve raporun dipnotunda
kullaniciya durustce gosterilir.
"""

from __future__ import annotations

import io
import logging

from app.documents.types import AyristirilmisBelge, Tablo

logger = logging.getLogger(__name__)

#: Kabul edilen uzantilar. Urun karari: YALNIZCA pdf ve excel (+ gorseller
#: ayri yoldan). Word/CSV/metin bilincli olarak DISARIDA - her format kendi
#: ayristirma tuzaklarini getirir ve test edilmeden acilmamalidir.
PDF_UZANTILARI = {".pdf"}
EXCEL_UZANTILARI = {".xlsx", ".xlsm", ".xls"}
GORSEL_UZANTILARI = {".png", ".jpg", ".jpeg", ".webp"}

#: Tek bir tablodan alinacak azami satir. Bir Excel sayfasi 100.000 satir
#: olabilir; tamami ne LLM baglamina ne de PDF'e sigar.
AZAMI_TABLO_SATIRI = 60

#: Excel'de taranacak azami sayfa (sheet).
AZAMI_EXCEL_SAYFASI = 10


class BelgeAyristirmaHatasi(Exception):
    """Dosya HIC ayristirilamadi - bu noktadan sonra rapor uretilemez."""


def belge_turu(dosya_adi: str) -> str:
    """Uzantidan belge turunu cikarir; desteklenmiyorsa firlatir.

    Uzantiya bakilir, ICERIGE degil: kullanicinin yukledigi dosyanin gercek
    turunu dogrulamak (magic byte) API katmaninin isidir - burada amac boru
    hattini dogru dala yonlendirmek.
    """
    nokta = dosya_adi.lower().rfind(".")
    uzanti = dosya_adi[nokta:].lower() if nokta >= 0 else ""

    if uzanti in PDF_UZANTILARI:
        return "pdf"
    if uzanti in EXCEL_UZANTILARI:
        return "excel"
    if uzanti in GORSEL_UZANTILARI:
        return "gorsel"

    raise BelgeAyristirmaHatasi(
        f"'{uzanti or dosya_adi}' desteklenmiyor. Yalnizca PDF, Excel (.xlsx/.xls) "
        "ve gorsel (.png/.jpg/.webp) dosyalari analiz edilebilir."
    )


def _temiz_hucre(deger) -> str:
    """Hucreyi PDF'e ve LLM'e guvenli tek satirlik metne cevirir.

    `None` -> "" : tabloda "None" yazisi gormek kullaniciyi yanlis yonlendirir.
    Satir sonlari bosluga cevrilir: ReportLab tablo hucresinde ham `\\n`
    satir yuksekligini bozar.
    """
    if deger is None:
        return ""
    metin = str(deger).strip().replace("\n", " ").replace("\r", " ")
    return " ".join(metin.split())


def _tablo_olustur(ham_satirlar: list, kaynak: str) -> Tablo | None:
    """Ham satir listesini `Tablo`'ya cevirir; ise yaramazsa `None`.

    Ilk satir BASLIK SAYILIR ancak yalnizca hucrelerinin cogu doluysa -
    aksi halde bassiz tablo uretilir. Yanlis baslik varsaymak, kullaniciya
    "Ocak" sutununu "Toplam" diye gostermek gibi somut bir hataya yol acar.
    """
    satirlar = [[_temiz_hucre(h) for h in satir] for satir in ham_satirlar if satir]
    satirlar = [s for s in satirlar if any(h for h in s)]
    if len(satirlar) < 2:
        return None

    ilk = satirlar[0]
    ilk_dolu_oran = sum(1 for h in ilk if h) / max(len(ilk), 1)
    if ilk_dolu_oran >= 0.6:
        return Tablo(basliklar=ilk, satirlar=satirlar[1 : AZAMI_TABLO_SATIRI + 1], kaynak=kaynak)
    return Tablo(basliklar=[], satirlar=satirlar[:AZAMI_TABLO_SATIRI], kaynak=kaynak)


def pdf_ayristir(icerik: bytes, dosya_adi: str) -> AyristirilmisBelge:
    """PDF'ten metin ve tablolari cikarir (pdfplumber)."""
    import pdfplumber

    metinler: list[str] = []
    tablolar: list[Tablo] = []
    uyarilar: list[str] = []
    metinsiz_sayfa = 0

    try:
        with pdfplumber.open(io.BytesIO(icerik)) as pdf:
            sayfa_sayisi = len(pdf.pages)
            for indeks, sayfa in enumerate(pdf.pages, start=1):
                try:
                    sayfa_metni = sayfa.extract_text() or ""
                except Exception:  # noqa: BLE001 - tek sayfa tum belgeyi dusurmemeli
                    logger.warning("PDF sayfasi okunamadi", extra={"sayfa": indeks})
                    sayfa_metni = ""

                if sayfa_metni.strip():
                    metinler.append(sayfa_metni)
                else:
                    # Taranmis (goruntu) sayfa: metin katmani yoktur. Sessizce
                    # atlanirsa kullanici raporun neden eksik oldugunu anlamaz.
                    metinsiz_sayfa += 1

                try:
                    for tablo_indeks, ham in enumerate(sayfa.extract_tables() or [], start=1):
                        tablo = _tablo_olustur(ham, f"sayfa {indeks} · tablo {tablo_indeks}")
                        if tablo:
                            tablolar.append(tablo)
                except Exception:  # noqa: BLE001
                    logger.warning("PDF tablosu okunamadi", extra={"sayfa": indeks})
    except Exception as hata:  # noqa: BLE001
        raise BelgeAyristirmaHatasi(
            f"PDF acilamadi (bozuk veya parola korumali olabilir): {hata}"
        ) from hata

    if metinsiz_sayfa:
        uyarilar.append(
            f"{sayfa_sayisi} sayfanin {metinsiz_sayfa} tanesi taranmis goruntu oldugu icin "
            "metni okunamadi; analiz kalan sayfalara dayaniyor."
        )
    if not metinler and not tablolar:
        raise BelgeAyristirmaHatasi(
            "PDF'ten hic metin veya tablo cikarilamadi. Belge tamamen taranmis "
            "goruntu olabilir; bu durumda gorsel olarak yukleyebilirsiniz."
        )

    return AyristirilmisBelge(
        tur="pdf",
        dosya_adi=dosya_adi,
        metin="\n\n".join(metinler),
        tablolar=tablolar,
        sayfa_sayisi=sayfa_sayisi,
        uyarilar=uyarilar,
    )


def excel_ayristir(icerik: bytes, dosya_adi: str) -> AyristirilmisBelge:
    """Excel sayfalarini tablolara cevirir (openpyxl).

    `data_only=True` KRITIK: formul iceren hucrelerde aksi halde `=B2*1.2`
    metni okunur, hesaplanan deger degil. Kullanici raporda formul gormemeli.
    """
    from openpyxl import load_workbook

    tablolar: list[Tablo] = []
    uyarilar: list[str] = []

    try:
        calisma_kitabi = load_workbook(io.BytesIO(icerik), data_only=True, read_only=True)
    except Exception as hata:  # noqa: BLE001
        raise BelgeAyristirmaHatasi(
            f"Excel dosyasi acilamadi (bozuk veya desteklenmeyen surum): {hata}"
        ) from hata

    try:
        sayfa_adlari = calisma_kitabi.sheetnames
        if len(sayfa_adlari) > AZAMI_EXCEL_SAYFASI:
            uyarilar.append(
                f"Dosyada {len(sayfa_adlari)} sayfa var; ilk {AZAMI_EXCEL_SAYFASI} "
                "sayfa analiz edildi."
            )

        for sayfa_adi in sayfa_adlari[:AZAMI_EXCEL_SAYFASI]:
            sayfa = calisma_kitabi[sayfa_adi]
            ham = []
            for satir in sayfa.iter_rows(values_only=True):
                ham.append(list(satir))
                # +1: baslik satiri tabloda satirdan sayilmaz.
                if len(ham) > AZAMI_TABLO_SATIRI + 1:
                    uyarilar.append(
                        f"'{sayfa_adi}' sayfasi cok uzun; ilk {AZAMI_TABLO_SATIRI} "
                        "satir analiz edildi."
                    )
                    break

            tablo = _tablo_olustur(ham, f"{sayfa_adi} (Excel)")
            if tablo:
                tablolar.append(tablo)
    finally:
        # read_only modda dosya tanitici acik kalir; kapatilmazsa Windows'ta
        # gecici dosya silinemez.
        calisma_kitabi.close()

    if not tablolar:
        raise BelgeAyristirmaHatasi(
            "Excel dosyasinda analiz edilebilecek dolu bir tablo bulunamadi."
        )

    return AyristirilmisBelge(
        tur="excel",
        dosya_adi=dosya_adi,
        metin="",
        tablolar=tablolar,
        sayfa_sayisi=len(tablolar),
        uyarilar=uyarilar,
    )


def ayristir(icerik: bytes, dosya_adi: str) -> AyristirilmisBelge:
    """Tur tespiti + dogru ayristiriciya yonlendirme (gorsel HARIC).

    Gorseller burada islenmez: onlar metin ayristirmasi degil GORME modeli
    ister ve ayri bir modulde (`vision.py`) ele alinir.
    """
    tur = belge_turu(dosya_adi)
    if tur == "pdf":
        return pdf_ayristir(icerik, dosya_adi)
    if tur == "excel":
        return excel_ayristir(icerik, dosya_adi)
    raise BelgeAyristirmaHatasi(
        "Gorseller bu fonksiyonla ayristirilmaz; `vision.gorseli_coz` kullanilmali."
    )
